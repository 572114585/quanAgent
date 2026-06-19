import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "AI新闻汇总"

header_font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
content_font = Font(name="Arial", size=10)
content_alignment = Alignment(vertical="center", wrap_text=True)
center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

headers = ["序号", "日期", "新闻标题", "新闻摘要", "来源/领域"]
col_widths = [6, 14, 35, 60, 18]

for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
    cell = ws.cell(row=1, column=col_idx, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border
    ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = w

news_data = [
    [1, "2026-06-08", "苹果WWDC 2026发布Siri AI与Apple Intelligence平台",
     "苹果在WWDC 2026上发布了一系列重磅更新，包括备受期待的Siri AI、全新的Apple Intelligence平台、儿童安全工具升级以及照片应用革新等十大重点功能。",
     "科技巨头"],
    [2, "2026-06-17", "美国出口禁令限制Anthropic AI模型，震动全球AI行业",
     "特朗普政府以安全为由，限制海外访问Anthropic的最新AI模型，迫使Anthropic在全球范围内关闭部分服务。CEO在G7峰会就此磋商。",
     "政策/监管"],
    [3, "2026-06-12", "Visa将AI接入ChatGPT——AI现在可以直接花你的钱",
     "Visa将支付功能直接接入ChatGPT，使AI代理能够代表用户执行支付操作。标志着AI从信息处理向金融行动的实质性跨越。",
     "金融科技"],
    [4, "2026-05-28", "Google全球发布Flow AI Agent",
     "Google推出Flow AI Agent面向全球用户开放。该AI代理可深度集成Google工作空间，帮助用户自动化完成复杂的工作流程任务。",
     "科技巨头"],
    [5, "2026-06-19", "VivaTech 2026：富士康与英伟达押注法国成为欧洲AI中心",
     "在巴黎VivaTech 2026大会上，富士康与英伟达宣布加大对法国的投资布局，将法国打造为欧洲AI基础设施中心。",
     "产业动态"],
    [6, "2026-06-16", "AI解码植物DNA'开关'，精准预测基因调控",
     "研究人员利用AI模型解码植物DNA中的调控区域，能够预测单个碱基变化对开花时间等性状的影响，为精准农业育种提供新工具。",
     "科学研究"],
    [7, "2026-06-11", "理想汽车CEO李想回应质疑，预告AI战略发布",
     "李想公开回应'车企跨界AI是不务正业'的质疑，预告将在LivisDay发布会上揭晓多项AI战略规划与技术细节。",
     "智能汽车"],
    [8, "2026-03-01", "越南首部《人工智能法》正式生效",
     "越南首部《人工智能法》生效，强调确保人类对AI系统所有决策和行为的控制与干预能力，为AI治理提供法律框架。",
     "政策/监管"],
    [9, "2026-06-19", "AI引发大规模裁员：Q1超5万人因AI失业",
     "2026年Q1报告显示，AI已成为企业裁员首要驱动力。Meta、亚马逊、Oracle等因引入AI大幅缩减人力成本，引发社会关注。",
     "社会影响"],
    [10, "2026-06-19", "欧洲新AI模型'W'崭露头角，挑战主流AI格局",
     "欧洲最新AI模型'W'被业界认为具备X Factor，有望成为ChatGPT、Claude之外的有力替代方案，在VivaTech 2026上受广泛关注。",
     "产业动态"],
]

even_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
odd_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

for row_idx, row_data in enumerate(news_data, 2):
    fill = even_fill if row_idx % 2 == 0 else odd_fill
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = content_font
        cell.border = thin_border
        if col_idx in (1, 2):
            cell.alignment = center_alignment
        else:
            cell.alignment = content_alignment
        cell.fill = fill

ws.row_dimensions[1].height = 30
for i in range(2, len(news_data) + 2):
    ws.row_dimensions[i].height = 50

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:E{len(news_data) + 1}"

output_path = "/workspace/AI新闻汇总_2026年6月.xlsx"
wb.save(output_path)
print(f"文件已保存至: {output_path}")
print(f"共收录 {len(news_data)} 条新闻")
