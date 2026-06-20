#!/usr/bin/env python3
"""Create a new Excel workbook with a title, headers, and rows.

Part of the excel-xlsx skill. Invoked by the agent via the `execute` tool:

    python skills/excel-xlsx/scripts/create.py \
        --out output/report.xlsx \
        --sheet Sales \
        --title "2026 Q2 Sales" \
        --headers "Product,Quantity,Unit Price,Total" \
        --rows '["Widget,10,9.9,=B2*C2"]' \
        --text-cols A

Design notes (these implement the SKILL.md Key Rules by default):
- Cells starting with `=` are written as formulas (never precomputed values).
- `--text-cols` forces a column to text so IDs / phones / ZIP codes keep
  leading zeros and are not truncated.
- Output is printed as a concise summary the agent can relay to the user.
"""
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

import openpyxl


def _parse_csv_row(raw: str) -> list[str]:
    """Split a single comma-separated row string into cell values.

    We do not use the csv module here because rows arrive as a single JSON
    string already; callers control the splitting via this simple rule.
    """
    return [cell.strip() for cell in raw.split(",")]


def _coerce_value(raw: str, text_cols: set[str], col_letter: str):
    """Convert a raw string cell into the value to store.

    - `=...` → formula (string, openpyxl keeps it as a formula)
    - column in text_cols → str (preserves leading zeros / long IDs)
    - looks like int/float → numeric
    - otherwise → str
    """
    if raw == "":
        return None
    if raw.startswith("="):
        return raw  # openpyxl treats a leading '=' as a formula
    if col_letter in text_cols:
        return str(raw)
    # try int
    try:
        if raw.lstrip("-+").isdigit():
            return int(raw)
    except Exception:
        pass
    # try float
    try:
        f = float(raw)
        # avoid converting things like "nan"/"inf" which float() accepts
        if raw.lower() in {"nan", "inf", "-inf", "+inf"}:
            return raw
        return f
    except ValueError:
        return raw


def col_letter(index: int) -> str:
    """0-based index → Excel column letter (0 -> A)."""
    result = ""
    n = index
    while True:
        n, rem = divmod(n, 26)
        result = chr(ord("A") + rem) + result
        if n == 0:
            break
        n -= 1
    return result


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Create a new Excel workbook (excel-xlsx skill).",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="Example:\n"
        "  create.py --out output/r.xlsx --headers 'A,B' "
        "--rows '[\"1,2\",\"3,4\"]' --text-cols A",
    )
    parser.add_argument("--out", required=True, help="Output .xlsx path (workspace-relative, e.g. output/r.xlsx).")
    parser.add_argument("--sheet", default="Sheet1", help="Worksheet name (default: Sheet1).")
    parser.add_argument("--title", default=None, help="Optional title written to A1 (row 1).")
    parser.add_argument("--headers", default=None, help="Comma-separated header values.")
    parser.add_argument(
        "--rows",
        default=None,
        help='JSON array of strings; each string is a comma-separated row. '
        'e.g. \'["a,1,2.5","b,3,=B2+1"]\'.',
    )
    parser.add_argument(
        "--text-cols",
        default="",
        help="Comma-separated column letters to force as text (e.g. 'A,C').",
    )
    args = parser.parse_args()

    out_path = _resolve_out_path(args.out)

    text_cols = {c.strip().upper() for c in args.text_cols.split(",") if c.strip()}

    headers = _parse_csv_row(args.headers) if args.headers else []

    rows_raw: list[str] = []
    if args.rows:
        try:
            parsed = json.loads(args.rows)
        except json.JSONDecodeError as e:
            print(f"ERROR: --rows is not valid JSON: {e}", file=sys.stderr)
            return 2
        if not isinstance(parsed, list):
            print("ERROR: --rows must be a JSON array of strings.", file=sys.stderr)
            return 2
        rows_raw = [str(r) for r in parsed]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = args.sheet

    current_row = 1
    if args.title:
        ws.cell(row=current_row, column=1, value=args.title)
        current_row += 1
        # leave a blank row after the title for readability
        current_row += 1

    if headers:
        for c_idx, h in enumerate(headers, start=1):
            ws.cell(row=current_row, column=c_idx, value=h)
        current_row += 1

    data_start_row = current_row
    n_rows = 0
    for row_str in rows_raw:
        cells = _parse_csv_row(row_str)
        for c_idx, raw_val in enumerate(cells, start=1):
            letter = col_letter(c_idx - 1)
            value = _coerce_value(raw_val, text_cols, letter)
            if value is not None:
                ws.cell(row=current_row, column=c_idx, value=value)
        current_row += 1
        n_rows += 1

    # Give header + title a bold style using named-style-friendly direct format.
    if args.title:
        ws["A1"].font = openpyxl.styles.Font(bold=True, size=14)
    if headers:
        header_row = data_start_row - 1
        for c_idx in range(1, len(headers) + 1):
            ws.cell(row=header_row, column=c_idx).font = openpyxl.styles.Font(bold=True)

    # Auto-size columns a little so the file isn't ugly on open.
    for col in ws.columns:
        letter = col[0].column_letter
        max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[letter].width = min(max(max_len + 2, 8), 50)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))

    # Concise summary for the agent to relay.
    formula_count = 0
    for row in ws.iter_rows():
        for c in row:
            if isinstance(c.value, str) and c.value.startswith("="):
                formula_count += 1

    print(f"OK created: {out_path}")
    print(f"  sheet: {args.sheet}")
    print(f"  rows: {ws.max_row}, cols: {ws.max_column}")
    print(f"  data rows: {n_rows}")
    print(f"  formulas: {formula_count}")
    if text_cols:
        print(f"  text columns: {', '.join(sorted(text_cols))}")
    return 0


def _resolve_out_path(out_arg: str) -> Path:
    """Resolve the --out argument to a real filesystem path.

    The agent passes workspace-relative paths like ``output/r.xlsx``. The
    _SkillsShellBackend already rewrites virtual paths so by the time this
    script runs the cwd is the workspace root and the path is already relative.
    For safety we also handle a leading slash (``/output/r.xlsx``) by stripping
    it, and keep absolute paths as-is.
    """
    p = Path(out_arg)
    if p.exists() or p.is_absolute() and len(p.parts) > 1 and not out_arg.startswith("/"):
        return p
    # legacy virtual path: strip leading slash, resolve under cwd
    if out_arg.startswith("/"):
        return Path(out_arg.lstrip("/"))
    return p


if __name__ == "__main__":
    raise SystemExit(main())
