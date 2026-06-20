#!/usr/bin/env python3
"""Inspect (read back) an Excel workbook for verification.

Part of the excel-xlsx skill. Read-only — never modifies the file.

    python skills/excel-xlsx/scripts/view.py --file output/report.xlsx
    python skills/excel-xlsx/scripts/view.py --file output/report.xlsx --show-formulas --max-rows 20

NOTE: this file is intentionally named `view.py`, not `inspect.py`, to avoid
shadowing the Python standard-library `inspect` module (which numpy and others
import). Because the script's own directory is sys.path[0], a stdlib name here
would break imports transitively.
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

import openpyxl


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Inspect an Excel workbook (excel-xlsx skill, read-only).",
    )
    parser.add_argument("--file", required=True, help="Workbook path to inspect (workspace-relative).")
    parser.add_argument("--sheet", default=None, help="Worksheet to show (default: first/active).")
    parser.add_argument("--max-rows", type=int, default=10, help="Max data rows to print (default 10).")
    parser.add_argument(
        "--show-formulas",
        action="store_true",
        help="Also print a formula audit (cells starting with '=').",
    )
    parser.add_argument("--list-sheets", action="store_true", help="Only list sheet names and exit.")
    args = parser.parse_args()

    file_path = _resolve_path(args.file)
    if not file_path.exists():
        print(f"ERROR: file not found: {file_path}", file=sys.stderr)
        return 2

    # Load non-read-only so max_row/max_column are reliable across openpyxl versions.
    wb = openpyxl.load_workbook(str(file_path), data_only=True)

    print(f"file: {file_path}")
    print(f"sheets: {wb.sheetnames}")

    if args.list_sheets:
        return 0

    ws = wb[args.sheet] if args.sheet else wb[wb.sheetnames[0]]
    print(f"sheet shown: {ws.title}")
    print(f"rows: {ws.max_row}, cols: {ws.max_column}")
    print()

    limit = max(0, args.max_rows)
    printed = 0
    for row in ws.iter_rows(values_only=True):
        if printed >= limit:
            remaining = (ws.max_row or 0) - printed
            if remaining > 0:
                print(f"... ({remaining} more rows)")
            break
        cells = ["" if v is None else _short(v) for v in row]
        print("  | ".join(cells))
        printed += 1

    if args.show_formulas:
        # Separate load with data_only=False to read the formula strings.
        wb2 = openpyxl.load_workbook(str(file_path), data_only=False)
        ws2 = wb2[ws.title]
        formula_cells = []
        for r_idx, row in enumerate(ws2.iter_rows(), start=1):
            for c_idx, cell in enumerate(row, start=1):
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formula_cells.append(f"{_col_letter(c_idx - 1)}{r_idx}={cell.value}")
        print()
        print(f"formulas ({len(formula_cells)}):")
        for line in formula_cells[:30]:
            print(f"  {line}")
        if len(formula_cells) > 30:
            print(f"  ... {len(formula_cells) - 30} more")

    return 0


def _short(v) -> str:
    s = str(v)
    return s if len(s) <= 40 else s[:37] + "..."


def _col_letter(index: int) -> str:
    result = ""
    n = index
    while True:
        n, rem = divmod(n, 26)
        result = chr(ord("A") + rem) + result
        if n == 0:
            break
        n -= 1
    return result


def _resolve_path(arg: str) -> Path:
    if arg.startswith("/"):
        return Path(arg.lstrip("/"))
    return Path(arg)


if __name__ == "__main__":
    raise SystemExit(main())
