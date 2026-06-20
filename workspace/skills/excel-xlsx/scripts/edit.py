#!/usr/bin/env python3
"""Edit an existing Excel workbook in place.

Part of the excel-xlsx skill. Invoked via the `execute` tool:

    python skills/excel-xlsx/scripts/edit.py \
        --file output/report.xlsx \
        --cell C5 --value '=B5*D5' \
        --cell D1 --value 'Notes' \
        --append-rows '["Widget,7,9.9,=B9*C9"]'

Safety rules (implement SKILL.md Key Rules):
- Loads with data_only=False so formulas are preserved on save.
- Never re-saves a workbook that was loaded with data_only=True.
"""
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

import openpyxl


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Edit an existing Excel workbook (excel-xlsx skill).",
    )
    parser.add_argument("--file", required=True, help="Workbook path to edit (workspace-relative).")
    parser.add_argument(
        "--cell",
        action="append",
        default=[],
        help="A target cell, e.g. 'C5'. Use with --value. Repeatable.",
    )
    parser.add_argument(
        "--value",
        action="append",
        default=[],
        help="Value for the corresponding --cell (same order). Repeatable.",
    )
    parser.add_argument(
        "--append-rows",
        default=None,
        help='JSON array of comma-separated row strings to append, e.g. \'["a,1","b,2"]\'.',
    )
    parser.add_argument("--sheet", default=None, help="Worksheet to edit (default: active sheet).")
    args = parser.parse_args()

    file_path = _resolve_path(args.file)
    if not file_path.exists():
        print(f"ERROR: file not found: {file_path}", file=sys.stderr)
        return 2

    # data_only=False: keep formulas. This is critical — loading with
    # data_only=True and saving would flatten formulas to static values.
    wb = openpyxl.load_workbook(str(file_path), data_only=False)
    ws = wb[args.sheet] if args.sheet else wb.active

    changes = 0

    if len(args.cell) != len(args.value):
        print(
            f"ERROR: {len(args.cell)} --cell args but {len(args.value)} --value args; they must match.",
            file=sys.stderr,
        )
        return 2

    for cell_ref, val in zip(args.cell, args.value):
        # formulas start with '='; store as-is. otherwise try numeric coercion
        # so the agent doesn't accidentally write "10" as text.
        value = _coerce(val)
        ws[cell_ref] = value
        changes += 1

    if args.append_rows is not None:
        try:
            new_rows = json.loads(args.append_rows)
        except json.JSONDecodeError as e:
            print(f"ERROR: --append-rows is not valid JSON: {e}", file=sys.stderr)
            return 2
        if not isinstance(new_rows, list):
            print("ERROR: --append-rows must be a JSON array of strings.", file=sys.stderr)
            return 2
        # find the next empty row after the last used row
        start_row = (ws.max_row or 0) + 1
        for r_idx, row_str in enumerate(new_rows):
            cells = [c.strip() for c in str(row_str).split(",")]
            for c_idx, raw in enumerate(cells, start=1):
                ws.cell(row=start_row + r_idx, column=c_idx, value=_coerce(raw))
            changes += 1

    wb.save(str(file_path))
    print(f"OK edited: {file_path}")
    print(f"  sheet: {ws.title}")
    print(f"  changes: {changes}")
    print(f"  rows now: {ws.max_row}, cols: {ws.max_column}")
    return 0


def _coerce(raw: str):
    """Coerce a string into int/float/str for storage. Keep leading '=' as formula."""
    if raw == "":
        return None
    if raw.startswith("="):
        return raw
    try:
        if raw.lstrip("-+").isdigit():
            return int(raw)
    except Exception:
        pass
    try:
        f = float(raw)
        if raw.lower() in {"nan", "inf", "-inf", "+inf"}:
            return raw
        return f
    except ValueError:
        return raw


def _resolve_path(arg: str) -> Path:
    if arg.startswith("/"):
        return Path(arg.lstrip("/"))
    return Path(arg)


if __name__ == "__main__":
    raise SystemExit(main())
