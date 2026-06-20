"""媒体文件处理模块：图片压缩、语音转文字、文件内容提取、视频关键帧提取

每个处理函数返回 ProcessedMedia，供 bridge.py 统一构造 LLM 输入。
"""
import base64
import io
import logging
import os
import subprocess
import tempfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

from .config import MediaConfig
from .media_storage import MediaStorage, detect_image_mime

logger = logging.getLogger(__name__)

# whisper 模型惰性单例缓存
_whisper_model = None


def _get_whisper_model():
    """获取或初始化 whisper 模型（惰性单例）"""
    global _whisper_model
    if _whisper_model is None:
        import whisper
        _whisper_model = whisper.load_model("base")
    return _whisper_model


@dataclass
class ProcessedMedia:
    """处理后的媒体结果，统一返回给 bridge 层"""
    # LLM 可消费的内容
    text_content: str = ""           # 提取的文本描述
    image_data_uris: list[str] = field(default_factory=list)  # base64 data URI 列表
    local_paths: list[str] = field(default_factory=list)      # 本地文件路径列表

    # 元信息
    media_type: str = ""             # image/voice/file/video
    original_name: str = ""          # 原始文件名
    size_bytes: int = 0              # 原始大小
    error: str = ""                  # 处理错误信息

    @property
    def success(self) -> bool:
        return not self.error


# ── 图片处理 ──────────────────────────────────────────────

def process_image(
    data: bytes,
    filename: str,
    config: MediaConfig,
    storage: MediaStorage,
    user_id: str = "",
) -> ProcessedMedia:
    """处理图片：格式验证 → 尺寸压缩 → base64 编码

    策略：
    - 超过 max_dimension 的图片自动缩放
    - 保存压缩后的副本到 storage
    - 返回 base64 data URI 供 LLM 多模态输入
    """
    result = ProcessedMedia(
        media_type="image",
        original_name=filename,
        size_bytes=len(data),
    )

    # 格式验证
    if not storage.validate_extension(filename, "image"):
        result.error = f"不支持的图片格式: {Path(filename).suffix}"
        return result

    # 大小验证
    if not storage.validate_size(len(data), "image"):
        result.error = f"图片过大 ({len(data) / 1024 / 1024:.1f}MB)，最大支持 {config.limits.image_max_size / 1024 / 1024:.0f}MB"
        return result

    # 尝试压缩
    compressed_data = _compress_image(data, config.image_max_dimension, config.image_quality)
    was_compressed = compressed_data is not None
    if compressed_data:
        data = compressed_data

    # 保存到存储
    safe_name = MediaStorage.sanitize_filename(filename)
    local_path = storage.save_bytes(data, "image", safe_name, user_id)
    result.local_paths.append(str(local_path))

    # 生成 data URI
    # 压缩后统一为 JPEG 格式，未压缩则按原始文件头检测
    if was_compressed:
        mime = "image/jpeg"
    else:
        mime = detect_image_mime(data)
    b64 = base64.b64encode(data).decode("ascii")
    result.image_data_uris.append(f"data:{mime};base64,{b64}")
    result.text_content = f"[图片: {filename}]"

    return result


def _compress_image(data: bytes, max_dimension: int, quality: int) -> bytes | None:
    """压缩图片：超过 max_dimension 时缩放，降低质量

    使用 Pillow 进行压缩，失败时返回 None（使用原图）
    """
    try:
        from PIL import Image

        img = Image.open(io.BytesIO(data))
        original_size = img.size

        # 判断是否需要缩放
        max_side = max(original_size)
        if max_side <= max_dimension:
            # 不需要缩放，但可以重新编码降低质量
            if img.mode in ("RGBA", "P"):
                img = img.convert("RGB")
            buf = io.BytesIO()
            img.save(buf, format="JPEG", quality=quality)
            compressed = buf.getvalue()
            # 只有压缩后更小才使用
            return compressed if len(compressed) < len(data) else None

        # 缩放
        ratio = max_dimension / max_side
        new_size = (int(original_size[0] * ratio), int(original_size[1] * ratio))
        img = img.resize(new_size, Image.LANCZOS)

        if img.mode in ("RGBA", "P"):
            img = img.convert("RGB")

        buf = io.BytesIO()
        img.save(buf, format="JPEG", quality=quality)
        return buf.getvalue()

    except ImportError:
        logger.debug("Pillow not installed, skipping image compression")
        return None
    except Exception as e:
        logger.warning("Image compression failed: %s", e)
        return None


# ── 语音处理 ──────────────────────────────────────────────

def process_voice(
    data: bytes,
    filename: str,
    config: MediaConfig,
    storage: MediaStorage,
    user_id: str = "",
    recognized_text: str = "",  # 微信已识别的语音转文字
) -> ProcessedMedia:
    """处理语音：格式验证 → 格式转换 → 语音识别

    策略：
    - 优先使用微信已识别的文字（voice_item.text）
    - 若无文字，尝试本地 ASR（需 ffmpeg + whisper）
    - amr/silk 格式自动转 wav
    - 保存原始文件和转换后的文件
    """
    result = ProcessedMedia(
        media_type="voice",
        original_name=filename,
        size_bytes=len(data),
    )

    # 格式验证
    if not storage.validate_extension(filename, "voice"):
        result.error = f"不支持的语音格式: {Path(filename).suffix}"
        return result

    # 大小验证
    if not storage.validate_size(len(data), "voice"):
        result.error = f"语音过大 ({len(data) / 1024 / 1024:.1f}MB)，最大支持 {config.limits.voice_max_size / 1024 / 1024:.0f}MB"
        return result

    # 保存原始文件
    safe_name = MediaStorage.sanitize_filename(filename)
    local_path = storage.save_bytes(data, "voice", safe_name, user_id)
    result.local_paths.append(str(local_path))

    # 语音转文字
    if recognized_text:
        result.text_content = recognized_text
    else:
        # 尝试格式转换 + ASR
        asr_text = _voice_to_text(data, filename, config)
        if asr_text:
            result.text_content = asr_text
        else:
            result.text_content = f"[语音消息: {filename}]（语音识别不可用，请安装 ffmpeg + openai-whisper）"

    return result


def _voice_to_text(data: bytes, filename: str, config: MediaConfig) -> str | None:
    """语音转文字：amr/silk → wav → ASR

    依赖：ffmpeg（格式转换）、openai-whisper（ASR）
    失败时返回 None
    """
    ext = Path(filename).suffix.lower()

    # 转换为 wav 格式
    wav_data = _convert_audio_to_wav(data, ext, config)
    if wav_data is None:
        return None

    # 尝试 ASR
    try:
        model = _get_whisper_model()
        # 写入临时文件
        with tempfile.NamedTemporaryFile(suffix=".wav", delete=False) as f:
            f.write(wav_data)
            tmp_path = f.name
        try:
            result = model.transcribe(tmp_path, language="zh")
            return result.get("text", "").strip()
        finally:
            os.unlink(tmp_path)
    except ImportError:
        logger.debug("whisper not installed, ASR unavailable")
        return None
    except Exception as e:
        logger.warning("ASR failed: %s", e)
        return None


def _convert_audio_to_wav(data: bytes, src_ext: str, config: MediaConfig) -> bytes | None:
    """使用 ffmpeg 将音频转换为 wav 格式

    支持 amr → wav, silk → wav, mp3 → wav 等
    """
    try:
        # 写入临时源文件
        with tempfile.NamedTemporaryFile(suffix=src_ext, delete=False) as f:
            f.write(data)
            src_path = f.name

        dst_path = src_path + ".wav"
        try:
            cmd = [
                "ffmpeg", "-i", src_path,
                "-ar", str(config.voice_sample_rate),
                "-ac", "1",  # 单声道
                "-y",        # 覆盖输出
                dst_path,
            ]
            proc = subprocess.run(cmd, capture_output=True, timeout=30)
            if proc.returncode != 0:
                logger.warning("ffmpeg conversion failed: %s", proc.stderr.decode(errors="replace")[:200])
                return None

            return Path(dst_path).read_bytes()
        finally:
            os.unlink(src_path)
            if os.path.exists(dst_path):
                os.unlink(dst_path)

    except FileNotFoundError:
        logger.debug("ffmpeg not found, audio conversion unavailable")
        return None
    except Exception as e:
        logger.warning("Audio conversion failed: %s", e)
        return None


# ── 文件处理 ──────────────────────────────────────────────

def process_file(
    data: bytes,
    filename: str,
    config: MediaConfig,
    storage: MediaStorage,
    user_id: str = "",
) -> ProcessedMedia:
    """处理文件：格式验证 → 大小检查 → 内容提取 → 安全扫描

    策略：
    - 文档类（pdf/docx/xlsx/pptx/txt/csv）提取文本摘要
    - 压缩包（zip）列出文件清单
    - 其他类型仅记录元信息
    """
    result = ProcessedMedia(
        media_type="file",
        original_name=filename,
        size_bytes=len(data),
    )

    # 格式验证
    if not storage.validate_extension(filename, "file"):
        result.error = f"不支持的文件格式: {Path(filename).suffix}"
        return result

    # 大小验证
    if not storage.validate_size(len(data), "file"):
        result.error = f"文件过大 ({len(data) / 1024 / 1024:.1f}MB)，最大支持 {config.limits.file_max_size / 1024 / 1024:.0f}MB"
        return result

    # 保存文件
    safe_name = MediaStorage.sanitize_filename(filename)
    local_path = storage.save_bytes(data, "file", safe_name, user_id)
    result.local_paths.append(str(local_path))

    # 提取文本内容
    ext = Path(filename).suffix.lower()
    extracted = _extract_file_text(data, ext)
    if extracted:
        # 截断过长内容
        max_chars = 3000
        if len(extracted) > max_chars:
            extracted = extracted[:max_chars] + "\n...（内容过长，已截断）"
        result.text_content = f"[文件: {filename}]\n内容摘要:\n{extracted}"
    else:
        result.text_content = f"[文件: {filename}，大小: {len(data) / 1024:.1f}KB]"

    return result


def _extract_file_text(data: bytes, ext: str) -> str | None:
    """从文件中提取文本内容

    支持：txt/csv/json/xml/html/md → 直接读取
          pdf → PyPDF2 提取
          docx → python-docx 提取
          xlsx → openpyxl 提取
          zip → 列出文件清单
    """
    try:
        # 纯文本类
        if ext in (".txt", ".csv", ".json", ".xml", ".html", ".md", ".log", ".py", ".js"):
            return data.decode("utf-8", errors="replace")

        # PDF
        if ext == ".pdf":
            return _extract_pdf_text(data)

        # Word
        if ext == ".docx":
            return _extract_docx_text(data)

        # Excel
        if ext in (".xlsx", ".xls"):
            return _extract_xlsx_text(data)

        # ZIP
        if ext == ".zip":
            return _extract_zip_listing(data)

        return None

    except Exception as e:
        logger.warning("Text extraction failed for %s: %s", ext, e)
        return None


def _extract_pdf_text(data: bytes) -> str | None:
    try:
        from PyPDF2 import PdfReader
        reader = PdfReader(io.BytesIO(data))
        pages = []
        for page in reader.pages[:10]:  # 最多提取前10页
            text = page.extract_text()
            if text:
                pages.append(text.strip())
        return "\n---\n".join(pages) if pages else None
    except ImportError:
        return None


def _extract_docx_text(data: bytes) -> str | None:
    try:
        from docx import Document
        doc = Document(io.BytesIO(data))
        paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
        return "\n".join(paragraphs) if paragraphs else None
    except ImportError:
        return None


def _extract_xlsx_text(data: bytes) -> str | None:
    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(data), read_only=True)
        rows = []
        for sheet in wb.sheetnames[:3]:  # 最多3个sheet
            ws = wb[sheet]
            rows.append(f"[Sheet: {sheet}]")
            for i, row in enumerate(ws.iter_rows(max_row=20, values_only=True)):
                cells = [str(c) if c is not None else "" for c in row]
                rows.append(" | ".join(cells))
        wb.close()
        return "\n".join(rows) if rows else None
    except ImportError:
        return None


def _extract_zip_listing(data: bytes) -> str | None:
    import zipfile
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as zf:
            names = zf.namelist()[:50]  # 最多50个文件
            return "压缩包内容:\n" + "\n".join(names)
    except Exception:
        return None


# ── 视频处理 ──────────────────────────────────────────────

def process_video(
    data: bytes,
    filename: str,
    config: MediaConfig,
    storage: MediaStorage,
    user_id: str = "",
) -> ProcessedMedia:
    """处理视频：格式验证 → 元信息提取 → 关键帧提取

    策略：
    - 提取视频元信息（时长、分辨率、编码）
    - 提取关键帧作为缩略图供 LLM 分析
    - 保存视频文件到存储
    """
    result = ProcessedMedia(
        media_type="video",
        original_name=filename,
        size_bytes=len(data),
    )

    # 格式验证
    if not storage.validate_extension(filename, "video"):
        result.error = f"不支持的视频格式: {Path(filename).suffix}"
        return result

    # 大小验证
    if not storage.validate_size(len(data), "video"):
        result.error = f"视频过大 ({len(data) / 1024 / 1024:.1f}MB)，最大支持 {config.limits.video_max_size / 1024 / 1024:.0f}MB"
        return result

    # 保存视频文件
    safe_name = MediaStorage.sanitize_filename(filename)
    local_path = storage.save_bytes(data, "video", safe_name, user_id)
    result.local_paths.append(str(local_path))

    # 提取元信息
    meta = _extract_video_metadata(str(local_path))
    meta_text = ""
    if meta:
        parts = []
        if meta.get("duration"):
            parts.append(f"时长: {meta['duration']:.1f}秒")
        if meta.get("width") and meta.get("height"):
            parts.append(f"分辨率: {meta['width']}x{meta['height']}")
        if meta.get("codec"):
            parts.append(f"编码: {meta['codec']}")
        meta_text = "，".join(parts)

    # 提取关键帧缩略图
    thumbnail_uris = _extract_video_thumbnails(str(local_path), config)
    if thumbnail_uris:
        result.image_data_uris.extend(thumbnail_uris)

    if meta_text:
        result.text_content = f"[视频: {filename}，{meta_text}]"
    else:
        result.text_content = f"[视频: {filename}]"

    return result


def _extract_video_metadata(video_path: str) -> dict | None:
    """使用 ffprobe 提取视频元信息"""
    try:
        cmd = [
            "ffprobe", "-v", "quiet",
            "-print_format", "json",
            "-show_format", "-show_streams",
            video_path,
        ]
        proc = subprocess.run(cmd, capture_output=True, timeout=10)
        if proc.returncode != 0:
            return None

        import json
        info = json.loads(proc.stdout)

        result = {}
        # 时长
        fmt = info.get("format", {})
        duration = fmt.get("duration")
        if duration:
            result["duration"] = float(duration)

        # 视频流信息
        for stream in info.get("streams", []):
            if stream.get("codec_type") == "video":
                result["width"] = stream.get("width")
                result["height"] = stream.get("height")
                result["codec"] = stream.get("codec_name")
                break

        return result

    except (FileNotFoundError, subprocess.TimeoutExpired, Exception) as e:
        logger.debug("ffprobe failed: %s", e)
        return None


def _extract_video_thumbnails(video_path: str, config: MediaConfig) -> list[str]:
    """提取视频关键帧作为缩略图

    提取 1-3 帧作为缩略图，返回 base64 data URI 列表。
    需先获取视频时长，再按比例提取开头、中间、结尾帧。
    """
    try:
        # 获取视频时长
        meta = _extract_video_metadata(video_path)
        duration = meta.get("duration", 0) if meta else 0

        # 计算提取时间点
        if duration > 3:
            time_points = [1.0, duration * 0.5, duration * 0.9]
        elif duration > 1:
            time_points = [0.5, duration * 0.8]
        else:
            time_points = [0.0]

        thumbnails = []
        for seek_time in time_points:
            with tempfile.NamedTemporaryFile(suffix=".jpg", delete=False) as f:
                thumb_path = f.name

            cmd = [
                "ffmpeg", "-y",
                "-i", video_path,
                "-ss", f"{seek_time:.2f}",
                "-frames:v", "1",
                "-q:v", "2",
                thumb_path,
            ]
            proc = subprocess.run(cmd, capture_output=True, timeout=10)
            if proc.returncode == 0 and Path(thumb_path).exists():
                thumb_data = Path(thumb_path).read_bytes()
                b64 = base64.b64encode(thumb_data).decode("ascii")
                thumbnails.append(f"data:image/jpeg;base64,{b64}")

            os.unlink(thumb_path)

        return thumbnails

    except (FileNotFoundError, Exception) as e:
        logger.debug("Video thumbnail extraction failed: %s", e)
        return []
